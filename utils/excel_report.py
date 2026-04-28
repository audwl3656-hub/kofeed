"""결과보고서 Excel 생성 유틸리티."""
import io
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

from utils.config import get_component_from_col, get_sample_from_col


# ── 스타일 헬퍼 ──────────────────────────────────────────────
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FONT = Font(name="맑은 고딕", size=10)
_FONT_B = Font(name="맑은 고딕", size=10, bold=True)
_FONT_W = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
_FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FILL_SUBHD  = PatternFill(start_color="8EAADB", end_color="8EAADB", fill_type="solid")
_FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ALIGN_C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _z_fill(z):
    try:
        v = float(z)
        if math.isnan(v):
            return None
        if abs(v) <= 2:
            return _FILL_GREEN
        elif abs(v) <= 3:
            return _FILL_YELLOW
        else:
            return _FILL_RED
    except (TypeError, ValueError):
        return None


def _hdr(ws, r, c, value, sub=False, wide=False):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = _FONT_W
    cell.fill = _FILL_SUBHD if sub else _FILL_HEADER
    cell.alignment = _ALIGN_C
    cell.border = _BORDER


def _val(ws, r, c, value, bold=False, align="center"):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = _FONT_B if bold else _FONT
    cell.alignment = _ALIGN_C if align == "center" else _ALIGN_L
    cell.border = _BORDER
    return cell


def _auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)


# ── 결과보고서 Excel ─────────────────────────────────────────
def generate_excel_summary(
    df,
    z_all,
    z_method: dict,
    group_stats: dict,
    value_cols: list,
    inst_field: str,
    generated_at: str,
    samples: list,
    participant_map: dict = None,
    subtitle: str = "",
    period_배부: str = "",
    period_회신: str = "",
    period_보고서: str = "",
    sample_note: str = "",
    summary_text: str = "",
    sample_comp_text: dict = None,
    cfg=None,
) -> bytes:
    """결과보고서 Excel — 개요 / 통계요약 / 전체Z-score 시트."""
    wb = Workbook()

    _name_to_code = {v: k for k, v in (participant_map or {}).items()}
    raw_names = df[inst_field].fillna("").astype(str).tolist()
    inst_codes = [_name_to_code.get(n, n) for n in raw_names]

    # 유효 컬럼 (통계 있는 것만)
    non_nir = [c for c in value_cols if group_stats.get(c)]
    seen_comps: list = []
    comp_to_cols: dict = {}
    for col in non_nir:
        comp = get_component_from_col(col, samples) or col
        if comp not in seen_comps:
            seen_comps.append(comp)
        comp_to_cols.setdefault(comp, []).append(col)

    # ── 시트 1: 개요 ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "개요"
    meta = [
        ("보고서 제목", "한국사료협회 비교분析 결과"),
        ("부제",        subtitle),
        ("생성일시",    generated_at),
        ("참가기관 수", len(df)),
        ("시료 배부",   period_배부),
        ("분析 및 결과회신", period_회신),
        ("결과처리 및 보고서 작성", period_보고서),
        ("시료 주석",   sample_note),
    ]
    for r, (k, v) in enumerate(meta, 1):
        _val(ws1, r, 1, k, bold=True, align="left")
        _val(ws1, r, 2, v, align="left")
    if sample_comp_text:
        r = len(meta) + 2
        _val(ws1, r, 1, "사료별 분析항목", bold=True, align="left")
        for i, (s, txt) in enumerate(sample_comp_text.items()):
            _val(ws1, r + i, 2, f"{s}: {txt}" if txt else s, align="left")
    if summary_text:
        r2 = len(meta) + len(sample_comp_text or {}) + 3
        _val(ws1, r2, 1, "분析결과 요약", bold=True, align="left")
        for i, line in enumerate(summary_text.splitlines()):
            _val(ws1, r2 + i, 2, line, align="left")
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 60

    # ── 시트 2: 통계요약 ──────────────────────────────────────
    ws2 = wb.create_sheet("통계요약")
    hdrs2 = ["성분", "사료", "n", "평균", "중앙값", "표준편차", "CV(%)", "최솟값", "최댓값"]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 1, ci, h)
    r = 2
    for comp in seen_comps:
        for col in comp_to_cols.get(comp, []):
            st = group_stats.get(col, {})
            if not st:
                continue
            s_label = get_sample_from_col(col, samples) or "-"
            vals = pd.to_numeric(df[col], errors="coerce").dropna()
            row_data = [
                comp, s_label,
                int(st.get("n", len(vals))),
                round(float(st.get("mean", 0)), 4),
                round(float(st.get("median", 0)), 4),
                round(float(st.get("std", 0)), 4),
                round(float(st.get("cv", 0)), 2) if not math.isnan(st.get("cv", float("nan"))) else "N/A",
                round(float(vals.min()), 4) if not vals.empty else "",
                round(float(vals.max()), 4) if not vals.empty else "",
            ]
            for ci, v in enumerate(row_data, 1):
                _val(ws2, r, ci, v, align="center" if ci > 1 else "left")
            r += 1
    _auto_width(ws2)

    # ── 시트 3: 전체 Z-score ──────────────────────────────────
    ws3 = wb.create_sheet("전체 Z-score")
    # 헤더: 기관코드 | 성분1_사료A_값 | 성분1_사료A_Z | ...
    col_headers = ["기관코드"]
    col_keys = []   # (col, "val"|"z")
    for comp in seen_comps:
        for col in comp_to_cols.get(comp, []):
            s = get_sample_from_col(col, samples) or col
            col_headers += [f"{comp}_{s}_값", f"{comp}_{s}_Z"]
            col_keys    += [(col, "val"), (col, "z")]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws3, 1, ci, h)
    for ri, (idx, row) in enumerate(df.iterrows(), 2):
        _val(ws3, ri, 1, inst_codes[ri - 2], bold=True)
        for ci, (col, kind) in enumerate(col_keys, 2):
            if kind == "val":
                v = row.get(col, "")
                try:
                    v = round(float(v), 4) if v != "" and not pd.isna(v) else ""
                except (TypeError, ValueError):
                    v = ""
                _val(ws3, ri, ci, v)
            else:
                z = z_all.loc[idx, col]
                try:
                    z_val = round(float(z), 3) if not pd.isna(z) else "N/A"
                except (TypeError, ValueError):
                    z_val = "N/A"
                cell = _val(ws3, ri, ci, z_val)
                fill = _z_fill(z_val)
                if fill:
                    cell.fill = fill
    _auto_width(ws3, min_w=8, max_w=20)

    # ── 시트 4: 방법별 Z-score ────────────────────────────────
    ws4 = wb.create_sheet("방법별 Z-score")
    col_headers4 = ["기관코드"]
    col_keys4 = []
    for comp in seen_comps:
        for col in comp_to_cols.get(comp, []):
            s = get_sample_from_col(col, samples) or col
            col_headers4 += [f"{comp}_{s}_값", f"{comp}_{s}_Z방법별"]
            col_keys4    += [(col, "val"), (col, "zm")]
    for ci, h in enumerate(col_headers4, 1):
        _hdr(ws4, 1, ci, h)
    for ri, (idx, row) in enumerate(df.iterrows(), 2):
        _val(ws4, ri, 1, inst_codes[ri - 2], bold=True)
        for ci, (col, kind) in enumerate(col_keys4, 2):
            if kind == "val":
                v = row.get(col, "")
                try:
                    v = round(float(v), 4) if v != "" and not pd.isna(v) else ""
                except (TypeError, ValueError):
                    v = ""
                _val(ws4, ri, ci, v)
            else:
                zm = z_method.get(col, pd.Series())
                try:
                    z_val = round(float(zm.loc[idx]), 3) if not pd.isna(zm.loc[idx]) else "N/A"
                except (KeyError, TypeError, ValueError):
                    z_val = "N/A"
                cell = _val(ws4, ri, ci, z_val)
                fill = _z_fill(z_val)
                if fill:
                    cell.fill = fill
    _auto_width(ws4, min_w=8, max_w=20)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
